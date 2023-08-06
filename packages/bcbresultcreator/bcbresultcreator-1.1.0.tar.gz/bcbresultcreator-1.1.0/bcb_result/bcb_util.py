import json
import re
import openpyxl as xl


def read_json_to_dict(file_name):
    with open(file_name) as fh:
        return json.load(fh)


def json_to_dict(json_str):
    return json.load(json_str)


def read_mean_time(pattern, folder_name):
    result = read_json_to_dict(f'{folder_name}/statistics.json')
    for key, value in result.items():
        if re.match(pattern, key):
            for k, v in value.items():
                if k == "meanResTime":
                    return v


def create_excel_with_text(folder, text):
    try:
        wb = xl.Workbook()
        sheet = wb.create_sheet('result')
        for index, t in enumerate(text):
            sheet.cell(row=index+1, column=1).value = t[0]
            sheet.cell(index+1, 2).value = t[1]
        wb.save(f'{folder}\jmeter_results.xlsx')
        return True
    except Exception as e:
        print('Something went wrong please check your error', e)
        return False
