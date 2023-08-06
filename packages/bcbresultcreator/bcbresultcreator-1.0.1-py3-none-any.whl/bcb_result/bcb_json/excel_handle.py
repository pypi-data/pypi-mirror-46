import openpyxl as xl


def create_excel_with_text(folder, text):
    try:
        wb = xl.Workbook()
        sheet = wb.create_sheet('result')
        for index, t in enumerate(text):
            sheet.cell(row=index+1, column=1).value = t[0]
            sheet.cell(index+1, 2).value = t[1]
        wb.save(f'{folder}\jmeter_results.xlsx')
        return True
    except Exception as e:
        print('Something went wrong please check your error', e)
        return False
