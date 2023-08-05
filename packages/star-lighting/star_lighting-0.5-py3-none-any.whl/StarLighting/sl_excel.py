import os
from pprint import pprint
import xlrd
import xlwt
from openpyxl import load_workbook, Workbook


class Excel:
    def __init__(self, file=None, index=0):
        self.rowNum = 0
        self.colNum = 0
        self.data = list()
        if file:
            if not os.path.isfile(file):
                raise FileNotFoundError('文件{}不存在'.format(file))
            else:
                self.load(file, index)

    def load(self, file: str, index=0):
        """'
        file:   str, 表示要加载的文件
        index： (int,str), 表示要加载的sheet
        """
        if not os.path.isfile(file):
            raise FileNotFoundError('文件不存在或不是文件')
        try:
            end = file.rsplit('.', 1)[-1]
        except Exception as e:
            raise TypeError(f'文件格式不准确, {e}')

        if end == 'xls':
            self.xls_load_list(file, index)
        elif end in ('xlsx', 'xlsm', 'xltx', 'xltm'):
            self.xlsx_load_list(file, index)
        else:
            raise TypeError('文件格式不正确')

        print(f'文件{file}加载完毕～')

    def xlsx_load_list(self, file: str, index: int = 0):
        """

        加载xlsx格式的文件
        :param file:
        :param index:
        :return:
        """
        wb = load_workbook(filename=file)
        sheet_names = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(sheet_names[index])
        rows = ws.rows
        columns = ws.columns

        _list = []
        for i, row in enumerate(rows):
            rlist = []  # 某一行
            for j, column in enumerate(row):
                # _arr[i][j] = column.value
                rlist.append(column.value)
            _list.append(rlist)
        self.data = _list
        self.rowNum = ws.max_row
        self.colNum = ws.max_column

    def xls_load_list(self, file: str, index: int = 0):
        workbook = xlrd.open_workbook(file)
        sheets = workbook.sheets()
        if isinstance(index, int):
            if index < len(sheets):
                sheet = workbook.sheets()[index]
            else:
                raise IndexError(f'输入的页码{index}不存在')
        elif isinstance(index, str):
            sheet = workbook.sheet_by_name(index)
        else:
            raise TypeError(f'{index}参数错误～')
        print(f'正在加载文件{file}...')
        self.rowNum = sheet.nrows  # sheet行数
        self.colNum = sheet.ncols  # sheet列数

        # 获取所有单元格的内容
        _arr = []
        for i in range(self.rowNum):  # 第0行到 rowNum-1 行
            rowlist = []
            for j in range(self.colNum):  # 第0列到 rowNum-1 列
                rowlist.append(sheet.cell_value(i, j))
            _arr.append(rowlist)
        self.data = _arr

    def make(self, data, filename='name1.xls', sheetname='sheet1'):
        """生成一个excel文件"""
        end = filename.rsplit('.', 1)[-1]
        if end not in ('xls', 'xlsx', 'xlsm', 'xltx', 'xltm'):
            raise TypeError(f'文件格式不正确, 只支持xls, xlsx, xlsm, xltx, xltm格式')
        if end == 'xls':
            f = xlwt.Workbook()
            sheet1 = f.add_sheet(sheetname=sheetname, cell_overwrite_ok=True)
            if isinstance(data, (list, tuple)):
                for i in range(len(data)):
                    for j in range(len(data[i])):
                        sheet1.write(i, j, f'{data[i][j]}')
            else:
                raise TypeError(f'文件{data}类型错误')
            f.save(filename)
        elif end in ('xls', 'xlsx', 'xlsm', 'xltx', 'xltm'):
            wb = Workbook()
            ws = wb.active
            ws.title = sheetname
            for l1 in data:
                ws.append(l1)
            wb.save(filename=filename)
        else:
            raise TypeError(f'文件格式不正确, 只支持xls, xlsx, xlsm, xltx, xltm格式')

    def __str__(self):
        return str(vars(self))

    def print_list(self):
        pprint(self.data)





